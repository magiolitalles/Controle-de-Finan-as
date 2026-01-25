import os
from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes, ConversationHandler
import pandas as pd
import openpyxl
from datetime import datetime
from dotenv import load_dotenv

# Carregar vari√°veis de ambiente
load_dotenv()

# Caminho do arquivo Excel
ARQUIVO_EXCEL = "planilha_financeira.xlsx"

# Estados da conversa
DESCRICAO, VALOR, TIPO, METODO, CATEGORIA = range(5)

# Fun√ß√µes auxiliares
def carregar_configuracoes():
    """Carrega as configura√ß√µes da planilha"""
    try:
        df_config = pd.read_excel(ARQUIVO_EXCEL, sheet_name='Configura√ß√µes', header=None)
        
        necessidades = df_config.iloc[1:, 0].dropna().tolist()
        desejos = df_config.iloc[1:, 1].dropna().tolist()
        investimentos = df_config.iloc[1:, 2].dropna().tolist()
        metodos = df_config.iloc[1:, 3].dropna().tolist()
        tipos = df_config.iloc[1:, 4].dropna().tolist()
        
        todas_categorias = necessidades + desejos + investimentos
        
        return {
            'todas_categorias': todas_categorias,
            'metodos': metodos,
            'tipos': tipos
        }
    except Exception as e:
        print(f"Erro ao carregar configura√ß√µes: {e}")
        return None

def adicionar_lancamento(data, descricao, categoria, tipo, valor, metodo):
    """Adiciona um novo lan√ßamento √† planilha"""
    try:
        wb = openpyxl.load_workbook(ARQUIVO_EXCEL)
        ws = wb['Lan√ßamentos']
        
        ultima_linha = ws.max_row + 1
        
        ws[f'A{ultima_linha}'] = data
        ws[f'B{ultima_linha}'] = descricao
        ws[f'C{ultima_linha}'] = categoria
        ws[f'D{ultima_linha}'] = tipo
        ws[f'E{ultima_linha}'] = valor
        ws[f'F{ultima_linha}'] = metodo
        ws[f'G{ultima_linha}'] = 'Realizado'
        
        wb.save(ARQUIVO_EXCEL)
        wb.close()
        
        return True
    except Exception as e:
        print(f"Erro ao adicionar lan√ßamento: {e}")
        return False

def calcular_saldo():
    """Calcula o saldo realizado e a transcorrer"""
    try:
        df = pd.read_excel(ARQUIVO_EXCEL, sheet_name='Lan√ßamentos')
        
        if df.empty:
            return None
        
        hoje = pd.Timestamp(datetime.now().date())
        
        # Calcular o √∫ltimo dia do m√™s seguinte
        if hoje.month == 12:
            proximo_mes = hoje.replace(year=hoje.year + 1, month=1, day=1)
        else:
            proximo_mes = hoje.replace(month=hoje.month + 1, day=1)
        
        if proximo_mes.month == 12:
            ultimo_dia_mes_seguinte = proximo_mes.replace(year=proximo_mes.year + 1, month=1, day=1) - pd.Timedelta(days=1)
        else:
            ultimo_dia_mes_seguinte = proximo_mes.replace(month=proximo_mes.month + 1, day=1) - pd.Timedelta(days=1)
        
        # Realizados
        df_realizados = df[df['Data'].dt.date <= hoje.date()]
        receitas_realizadas = df_realizados[df_realizados['Tipo'] == 'Receita']['Valor'].sum()
        despesas_realizadas = df_realizados[df_realizados['Tipo'] == 'Despesa']['Valor'].sum()
        saldo_realizado = receitas_realizadas - despesas_realizadas
        
        # A transcorrer
        amanha = hoje + pd.Timedelta(days=1)
        df_transcorrer = df[
            (df['Data'].dt.date >= amanha.date()) & 
            (df['Data'].dt.date <= ultimo_dia_mes_seguinte.date())
        ]
        receitas_transcorrer = df_transcorrer[df_transcorrer['Tipo'] == 'Receita']['Valor'].sum()
        despesas_transcorrer = df_transcorrer[df_transcorrer['Tipo'] == 'Despesa']['Valor'].sum()
        
        # Pr√≥xima receita (agrupada por data)
        receitas_futuras = df[
            (df['Data'].dt.date >= amanha.date()) & 
            (df['Tipo'] == 'Receita')
        ]
        
        proxima_receita = None
        proxima_receita_data = None
        if not receitas_futuras.empty:
            # Agrupar por data e somar os valores
            receitas_agrupadas = receitas_futuras.groupby('Data')['Valor'].sum().reset_index()
            receitas_agrupadas = receitas_agrupadas.sort_values('Data')
            proxima_receita = receitas_agrupadas.iloc[0]['Valor']
            proxima_receita_data = receitas_agrupadas.iloc[0]['Data']
        
        # Pr√≥xima despesa (agrupada por data)
        despesas_futuras = df[
            (df['Data'].dt.date >= amanha.date()) & 
            (df['Tipo'] == 'Despesa')
        ]
        
        proxima_despesa = None
        proxima_despesa_data = None
        if not despesas_futuras.empty:
            # Agrupar por data e somar os valores
            despesas_agrupadas = despesas_futuras.groupby('Data')['Valor'].sum().reset_index()
            despesas_agrupadas = despesas_agrupadas.sort_values('Data')
            proxima_despesa = despesas_agrupadas.iloc[0]['Valor']
            proxima_despesa_data = despesas_agrupadas.iloc[0]['Data']
        
        # √öltimo lan√ßamento (mais recente at√© a data atual)
        df_ate_hoje = df[df['Data'].dt.date <= hoje.date()]
        df_ordenado = df_ate_hoje.sort_values('Data', ascending=False)
        ultimo_lancamento = None
        ultimo_lancamento_data = None
        ultimo_lancamento_descricao = None
        ultimo_lancamento_tipo = None
        ultimo_lancamento_valor = None
        
        if not df_ordenado.empty:
            ultimo = df_ordenado.iloc[0]
            ultimo_lancamento_data = ultimo['Data']
            ultimo_lancamento_descricao = ultimo['Descri√ß√£o']
            ultimo_lancamento_tipo = ultimo['Tipo']
            ultimo_lancamento_valor = ultimo['Valor']
        
        return {
            'saldo_realizado': saldo_realizado,
            'receitas_transcorrer': receitas_transcorrer,
            'despesas_transcorrer': despesas_transcorrer,
            'proxima_receita': proxima_receita,
            'proxima_receita_data': proxima_receita_data,
            'proxima_despesa': proxima_despesa,
            'proxima_despesa_data': proxima_despesa_data,
            'ultimo_lancamento_data': ultimo_lancamento_data,
            'ultimo_lancamento_descricao': ultimo_lancamento_descricao,
            'ultimo_lancamento_tipo': ultimo_lancamento_tipo,
            'ultimo_lancamento_valor': ultimo_lancamento_valor
        }
    except Exception as e:
        print(f"Erro ao calcular saldo: {e}")
        return None

# Comandos do bot
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Comando /start"""
    mensagem = (
        "ü§ñ *Bem-vindo ao Bot de Controle Financeiro!*\n\n"
        "Comandos dispon√≠veis:\n"
        "/novo - Adicionar novo lan√ßamento\n"
        "/saldo - Ver saldo e resumo\n"
        "/historico - Ver √∫ltimos 5 lan√ßamentos\n"
        "/cancelar - Cancelar opera√ß√£o atual\n"
        "/ajuda - Ver esta mensagem"
    )
    await update.message.reply_text(mensagem, parse_mode='Markdown')

async def ajuda(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Comando /ajuda"""
    await start(update, context)

async def historico(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Comando /historico - Mostra os √∫ltimos 5 lan√ßamentos j√° realizados"""
    try:
        df = pd.read_excel(ARQUIVO_EXCEL, sheet_name='Lan√ßamentos')
        
        if df.empty:
            await update.message.reply_text("üì≠ Nenhum lan√ßamento registrado ainda.")
            return
        
        # Filtrar apenas lan√ßamentos at√© hoje
        hoje = pd.Timestamp(datetime.now().date())
        df_realizados = df[df['Data'].dt.date <= hoje.date()]
        
        if df_realizados.empty:
            await update.message.reply_text("üì≠ Nenhum lan√ßamento realizado at√© hoje.")
            return
        
        # Ordenar por data decrescente e pegar os √∫ltimos 5
        df_ordenado = df_realizados.sort_values('Data', ascending=False).head(5)
        
        mensagem = "üìù *√öltimos 5 Lan√ßamentos:*\n\n"
        
        for idx, row in df_ordenado.iterrows():
            data_formatada = row['Data'].strftime('%d/%m/%Y')            
            mensagem += (
                f"*{row['Tipo']}*\n"
                f"üìù {row['Descri√ß√£o']}\n"
                f"üí∞ R$ {row['Valor']:,.2f}\n"
                f"üè∑Ô∏è {row['Categoria']}\n"
                f"üí≥ {row['M√©todo']}\n"
                f"üìÖ {data_formatada}\n\n"
            )
        
        await update.message.reply_text(mensagem, parse_mode='Markdown')
        
    except Exception as e:
        print(f"Erro ao buscar hist√≥rico: {e}")
        await update.message.reply_text("‚ùå Erro ao buscar hist√≥rico. Tente novamente.")

async def saldo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Comando /saldo"""
    resultado = calcular_saldo()
    
    if resultado is None:
        await update.message.reply_text("üì≠ Nenhum lan√ßamento registrado ainda.")
        return
    
    mensagem = (
        "üí∞ *Resumo Financeiro*\n\n"
        f"‚úÖ *Saldo:* R$ {resultado['saldo_realizado']:,.2f}\n\n"
        f"üìÖ *A Transcorrer:*\n"
        f"Receitas: R$ {resultado['receitas_transcorrer']:,.2f}\n"
        f"Despesas: R$ {resultado['despesas_transcorrer']:,.2f}\n\n"
    )
    
    # Adicionar informa√ß√£o do √∫ltimo lan√ßamento
    if resultado['ultimo_lancamento_data'] is not None:
        data_formatada = resultado['ultimo_lancamento_data'].strftime('%d/%m/%Y')
        mensagem += (
            f"üìù *√öltimo Lan√ßamento:*\n"
            f"{resultado['ultimo_lancamento_tipo']}: {resultado['ultimo_lancamento_descricao']}\n"
            f"Valor: R$ {resultado['ultimo_lancamento_valor']:,.2f}\n"
            f"Data: {data_formatada}\n\n"
        )
    else:
        mensagem += "üìù *√öltimo Lan√ßamento:* Nenhum lan√ßamento registrado\n\n"
    
    # Adicionar informa√ß√£o da pr√≥xima receita
    if resultado['proxima_receita'] is not None:
        data_formatada = resultado['proxima_receita_data'].strftime('%d/%m/%Y')
        mensagem += (
            f"üíµ *Pr√≥xima Receita:*\n"
            f"Valor: R$ {resultado['proxima_receita']:,.2f}\n"
            f"Data: {data_formatada}\n\n"
        )
    else:
        mensagem += "üíµ *Pr√≥xima Receita:* Nenhuma receita futura registrada\n\n"
    
    # Adicionar informa√ß√£o da pr√≥xima despesa
    if resultado['proxima_despesa'] is not None:
        data_formatada = resultado['proxima_despesa_data'].strftime('%d/%m/%Y')
        mensagem += (
            f"üí≥ *Pr√≥xima Despesa:*\n"
            f"Valor: R$ {resultado['proxima_despesa']:,.2f}\n"
            f"Data: {data_formatada}"
        )
    else:
        mensagem += "üí≥ *Pr√≥xima Despesa:* Nenhuma despesa futura registrada"
    
    await update.message.reply_text(mensagem, parse_mode='Markdown')

async def novo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Inicia o processo de adicionar novo lan√ßamento"""
    await update.message.reply_text(
        "üìù Vamos adicionar um novo lan√ßamento!\n\n"
        "Por favor, envie a *descri√ß√£o* do lan√ßamento:\n"
        "(Ex: Almo√ßo, Compras, Sal√°rio, etc.)\n\n"
        "Use /cancelar para cancelar.",
        parse_mode='Markdown'
    )
    return DESCRICAO

async def receber_descricao(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Recebe a descri√ß√£o"""
    context.user_data['descricao'] = update.message.text
    
    await update.message.reply_text(
        f"‚úÖ Descri√ß√£o: *{update.message.text}*\n\n"
        "üí∞ Agora, envie o *valor*:\n"
        "(Ex: 50.00 ou 50)",
        parse_mode='Markdown'
    )
    return VALOR

async def receber_valor(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Recebe o valor"""
    try:
        valor = float(update.message.text.replace(',', '.'))
        if valor <= 0:
            await update.message.reply_text("‚ö†Ô∏è O valor deve ser maior que zero. Tente novamente:")
            return VALOR
        
        context.user_data['valor'] = valor
        
        # Criar teclado com os tipos (fixos)
        tipos = ['Receita', 'Despesa']
        keyboard = [[tipo] for tipo in tipos]
        reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
        
        await update.message.reply_text(
            f"‚úÖ Valor: *R$ {valor:.2f}*\n\n"
            "üíµ Selecione o *tipo*:",
            reply_markup=reply_markup,
            parse_mode='Markdown'
        )
        return TIPO
    except ValueError:
        await update.message.reply_text("‚ö†Ô∏è Valor inv√°lido. Use n√∫meros (Ex: 50.00). Tente novamente:")
        return VALOR

async def receber_tipo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Recebe o tipo"""
    tipo = update.message.text
    
    # Validar se o tipo √© v√°lido
    if tipo not in ['Receita', 'Despesa']:
        await update.message.reply_text(
            "‚ö†Ô∏è Tipo inv√°lido! Selecione apenas *Receita* ou *Despesa*.",
            parse_mode='Markdown'
        )
        return TIPO
    
    context.user_data['tipo'] = tipo
    
    config = carregar_configuracoes()
    if config is None:
        await update.message.reply_text("‚ùå Erro ao carregar configura√ß√µes.")
        return ConversationHandler.END
    
    # Criar teclado com os m√©todos
    keyboard = [[metodo] for metodo in config['metodos']]
    reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
    
    await update.message.reply_text(
        f"‚úÖ Tipo: *{tipo}*\n\n"
        "üí≥ Selecione o *m√©todo de pagamento*:",
        reply_markup=reply_markup,
        parse_mode='Markdown'
    )
    return METODO

async def receber_metodo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Recebe o m√©todo"""
    context.user_data['metodo'] = update.message.text
    
    config = carregar_configuracoes()
    if config is None:
        await update.message.reply_text("‚ùå Erro ao carregar configura√ß√µes.")
        return ConversationHandler.END
    
    # Criar teclado com as categorias (2 por linha para facilitar)
    categorias = config['todas_categorias']
    keyboard = [categorias[i:i+2] for i in range(0, len(categorias), 2)]
    reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
    
    await update.message.reply_text(
        f"‚úÖ M√©todo: *{update.message.text}*\n\n"
        "üè∑Ô∏è Selecione a *categoria*:",
        reply_markup=reply_markup,
        parse_mode='Markdown'
    )
    return CATEGORIA

async def receber_categoria(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Recebe a categoria e finaliza o lan√ßamento"""
    context.user_data['categoria'] = update.message.text
    
    # Salvar o lan√ßamento
    data = datetime.now()
    descricao = context.user_data['descricao']
    valor = context.user_data['valor']
    tipo = context.user_data['tipo']
    metodo = context.user_data['metodo']
    categoria = context.user_data['categoria']
    
    if adicionar_lancamento(data, descricao, categoria, tipo, valor, metodo):
        mensagem = (
            "‚úÖ *Lan√ßamento adicionado com sucesso!*\n\n"
            f"üìù Descri√ß√£o: {descricao}\n"
            f"üí∞ Valor: R$ {valor:.2f}\n"
            f"üíµ Tipo: {tipo}\n"
            f"üí≥ M√©todo: {metodo}\n"
            f"üè∑Ô∏è Categoria: {categoria}\n"
            f"üìÖ Data: {data.strftime('%d/%m/%Y')}"
        )
        await update.message.reply_text(
            mensagem,
            reply_markup=ReplyKeyboardRemove(),
            parse_mode='Markdown'
        )
    else:
        await update.message.reply_text(
            "‚ùå Erro ao adicionar lan√ßamento. Tente novamente.",
            reply_markup=ReplyKeyboardRemove()
        )
    
    # Limpar dados do usu√°rio
    context.user_data.clear()
    
    return ConversationHandler.END

async def cancelar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Cancela a opera√ß√£o atual"""
    context.user_data.clear()
    await update.message.reply_text(
        "‚ùå Opera√ß√£o cancelada.",
        reply_markup=ReplyKeyboardRemove()
    )
    return ConversationHandler.END

def main():
    """Fun√ß√£o principal"""
    # Carregar token do arquivo .env
    TOKEN = os.getenv('TELEGRAM_BOT_TOKEN')
    
    if not TOKEN:
        print("‚ùå ERRO: Token do Telegram n√£o encontrado!")
        return
    
    # Criar a aplica√ß√£o
    application = Application.builder().token(TOKEN).build()
    
    # Handler de conversa para adicionar lan√ßamento
    conv_handler = ConversationHandler(
        entry_points=[CommandHandler('novo', novo)],
        states={
            DESCRICAO: [MessageHandler(filters.TEXT & ~filters.COMMAND, receber_descricao)],
            VALOR: [MessageHandler(filters.TEXT & ~filters.COMMAND, receber_valor)],
            TIPO: [MessageHandler(filters.TEXT & ~filters.COMMAND, receber_tipo)],
            METODO: [MessageHandler(filters.TEXT & ~filters.COMMAND, receber_metodo)],
            CATEGORIA: [MessageHandler(filters.TEXT & ~filters.COMMAND, receber_categoria)],
        },
        fallbacks=[CommandHandler('cancelar', cancelar)]
    )
    
    # Adicionar handlers
    application.add_handler(CommandHandler('start', start))
    application.add_handler(CommandHandler('ajuda', ajuda))
    application.add_handler(CommandHandler('saldo', saldo))
    application.add_handler(CommandHandler('historico', historico))
    application.add_handler(conv_handler)
    
    # Iniciar o bot
    print("Bot Telegram iniciado! Aguardando mensagens...")
    application.run_polling(allowed_updates=Update.ALL_TYPES)

if __name__ == '__main__':
    main()
