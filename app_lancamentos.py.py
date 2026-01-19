import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os
import warnings
from openpyxl import Workbook

# Ignorar avisos do openpyxl
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# Configura√ß√£o da p√°gina
st.set_page_config(
    page_title="Registro de Lan√ßamentos Financeiros",
    page_icon="üí∞",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Caminho do arquivo Excel
ARQUIVO_EXCEL = "planilha_financeira.xlsx"

# Fun√ß√£o para carregar categorias e m√©todos da planilha
@st.cache_data
def carregar_configuracoes():
    """Carrega as configura√ß√µes da aba Configura√ß√µes"""
    try:
        df_config = pd.read_excel(ARQUIVO_EXCEL, sheet_name='Configura√ß√µes', header=None)
        
        # Extrair categorias
        necessidades = df_config.iloc[1:, 0].dropna().tolist()
        desejos = df_config.iloc[1:, 1].dropna().tolist()
        investimentos = df_config.iloc[1:, 2].dropna().tolist()
        
        # Extrair m√©todos
        metodos = df_config.iloc[1:, 3].dropna().tolist()
        
        # Extrair tipos
        tipos = df_config.iloc[1:, 4].dropna().tolist()
        
        todas_categorias = necessidades + desejos + investimentos
        
        return {
            'necessidades': necessidades,
            'desejos': desejos,
            'investimentos': investimentos,
            'todas_categorias': todas_categorias,
            'metodos': metodos,
            'tipos': tipos
        }
    except Exception as e:
        st.error(f"Erro ao carregar configura√ß√µes: {e}")
        return None

# Fun√ß√£o para carregar lan√ßamentos existentes
def carregar_lancamentos():
    """Carrega os lan√ßamentos da aba Lan√ßamentos"""
    try:
        df = pd.read_excel(ARQUIVO_EXCEL, sheet_name='Lan√ßamentos')
        return df
    except Exception as e:
        st.error(f"Erro ao carregar lan√ßamentos: {e}")
        return pd.DataFrame()

# Fun√ß√£o para adicionar novo lan√ßamento
def adicionar_lancamento(data, descricao, categoria, tipo, valor, metodo, status='Realizado'):
    """Adiciona um novo lan√ßamento √† planilha"""
    try:
        # Carregar o arquivo Excel
        wb = openpyxl.load_workbook(ARQUIVO_EXCEL)
        ws = wb['Lan√ßamentos']
        
        # Encontrar a pr√≥xima linha vazia
        ultima_linha = ws.max_row + 1
        
        # Inserir dados
        ws[f'A{ultima_linha}'] = data
        ws[f'B{ultima_linha}'] = descricao
        ws[f'C{ultima_linha}'] = categoria
        ws[f'D{ultima_linha}'] = tipo
        ws[f'E{ultima_linha}'] = valor
        ws[f'F{ultima_linha}'] = metodo
        ws[f'G{ultima_linha}'] = status
        
        # Salvar o arquivo
        wb.save(ARQUIVO_EXCEL)
        wb.close()
        
        return True
    except Exception as e:
        st.error(f"Erro ao adicionar lan√ßamento: {e}")
        return False

# Estilo CSS personalizado
st.markdown("""
    <style>
    .main-header {
        text-align: center;
        color: #1F497D;
        margin-bottom: 30px;
    }
    .success-box {
        background-color: #d4edda;
        border: 1px solid #c3e6cb;
        color: #155724;
        padding: 12px;
        border-radius: 4px;
        margin-bottom: 20px;
    }
    .info-box {
        background-color: #d1ecf1;
        border: 1px solid #bee5eb;
        color: #0c5460;
        padding: 12px;
        border-radius: 4px;
        margin-bottom: 20px;
    }
    </style>
""", unsafe_allow_html=True)

# T√≠tulo
st.markdown("<h1 class='main-header'>üí∞ Registro de Lan√ßamentos Financeiros</h1>", unsafe_allow_html=True)

# Carregar configura√ß√µes
config = carregar_configuracoes()

if config is None:
    st.error("N√£o foi poss√≠vel carregar as configura√ß√µes. Verifique se o arquivo Excel est√° correto.")
    st.stop()

# Sidebar com informa√ß√µes
with st.sidebar:
    st.header("üìä Informa√ß√µes")
    st.info("Use este aplicativo para registrar seus gastos e receitas de forma simples e r√°pida.")
    
    # Mostrar √∫ltimos lan√ßamentos
    st.subheader("üìù √öltimos Lan√ßamentos")
    df_lancamentos = carregar_lancamentos()
    
    # Corrigir tipos de dados no DataFrame
    for col in df_lancamentos.select_dtypes(include=['object']).columns:
        df_lancamentos[col] = df_lancamentos[col].astype(str)
    
    if not df_lancamentos.empty:
        ultimos = df_lancamentos.tail(5)[['Data', 'Descri√ß√£o', 'Tipo', 'Valor']].copy()
        st.dataframe(ultimos, use_container_width=True)
    else:
        st.write("Nenhum lan√ßamento registrado ainda.")

# Abas principais
tab1, tab2, tab3 = st.tabs(["üì• Novo Lan√ßamento", "üìä Visualizar Lan√ßamentos", "üìà Resumo"])

# TAB 1: Novo Lan√ßamento
with tab1:
    st.subheader("Adicione um novo lan√ßamento")
    
    col1, col2 = st.columns(2)
    
    with col1:
        data = st.date_input(
            "üìÖ Data do Lan√ßamento",
            value=datetime.now(),
            help="Selecione a data do lan√ßamento"
        )
        
        descricao = st.text_input(
            "üìù Descri√ß√£o",
            placeholder="Ex: Almo√ßo, Compras, Sal√°rio...",
            help="Descreva brevemente o lan√ßamento"
        )
        
        categoria = st.selectbox(
            "üè∑Ô∏è Categoria",
            options=config['todas_categorias'],
            help="Selecione a categoria do lan√ßamento"
        )
    
    with col2:
        tipo = st.radio(
            "üíµ Tipo",
            options=config['tipos'],
            horizontal=True,
            help="Selecione se √© receita ou despesa"
        )
        
        valor = st.number_input(
            "üí∞ Valor",
            min_value=0.0,
            step=0.01,
            format="%.2f",
            help="Insira o valor do lan√ßamento"
        )
        
        metodo = st.selectbox(
            "üí≥ M√©todo de Pagamento",
            options=config['metodos'],
            help="Selecione o m√©todo de pagamento"
        )
    
    # Bot√£o para adicionar
    col_btn1, col_btn2, col_btn3 = st.columns([1, 1, 2])
    
    with col_btn1:
        if st.button("‚úÖ Adicionar Lan√ßamento", use_container_width=True):
            if not descricao:
                st.error("‚ö†Ô∏è Por favor, preencha a descri√ß√£o.")
            elif valor <= 0:
                st.error("‚ö†Ô∏è O valor deve ser maior que zero.")
            else:
                # Converter data para datetime
                data_datetime = pd.Timestamp(data)
                
                if adicionar_lancamento(data_datetime, descricao, categoria, tipo, valor, metodo):
                    st.markdown(
                        f"<div class='success-box'><strong>‚úÖ Sucesso!</strong> Lan√ßamento adicionado com sucesso!</div>",
                        unsafe_allow_html=True
                    )
                    # Limpar cache para atualizar dados
                    st.cache_data.clear()
                else:
                    st.error("‚ùå Erro ao adicionar lan√ßamento.")
    
    with col_btn2:
        if st.button("üîÑ Limpar Formul√°rio", use_container_width=True):
            st.rerun()
    
    # Resumo Financeiro Realizado vs A Transcorrer
    st.divider()
    st.subheader("üí∞ Resumo Financeiro")
    
    # Carregar lan√ßamentos
    df_resumo = carregar_lancamentos()
    
    if not df_resumo.empty:
        # Data atual
        hoje = pd.Timestamp(datetime.now().date())
        
        # Calcular o √∫ltimo dia do m√™s seguinte
        if hoje.month == 12:
            proximo_mes = hoje.replace(year=hoje.year + 1, month=1, day=1)
        else:
            proximo_mes = hoje.replace(month=hoje.month + 1, day=1)
        
        # √öltimo dia do m√™s seguinte
        if proximo_mes.month == 12:
            ultimo_dia_mes_seguinte = proximo_mes.replace(year=proximo_mes.year + 1, month=1, day=1) - pd.Timedelta(days=1)
        else:
            ultimo_dia_mes_seguinte = proximo_mes.replace(month=proximo_mes.month + 1, day=1) - pd.Timedelta(days=1)
        
        # Realizados: at√© hoje (inclusive)
        df_realizados = df_resumo[df_resumo['Data'].dt.date <= hoje.date()]
        receitas_realizadas = df_realizados[df_realizados['Tipo'] == 'Receita']['Valor'].sum()
        despesas_realizadas = df_realizados[df_realizados['Tipo'] == 'Despesa']['Valor'].sum()
        saldo_realizado = receitas_realizadas - despesas_realizadas
        
        # A transcorrer: de amanh√£ at√© o final do m√™s seguinte
        amanha = hoje + pd.Timedelta(days=1)
        df_transcorrer = df_resumo[
            (df_resumo['Data'].dt.date >= amanha.date()) & 
            (df_resumo['Data'].dt.date <= ultimo_dia_mes_seguinte.date())
        ]
        receitas_transcorrer = df_transcorrer[df_transcorrer['Tipo'] == 'Receita']['Valor'].sum()
        despesas_transcorrer = df_transcorrer[df_transcorrer['Tipo'] == 'Despesa']['Valor'].sum()
        
        # Exibir m√©tricas
        col_res1, col_res2, col_res3 = st.columns(3)
        
        with col_res1:
            st.metric("‚úÖ Saldo Realizado (at√© hoje)", f"R$ {saldo_realizado:,.2f}")
        
        with col_res2:
            st.metric("üìÖ Despesas a Transcorrer", f"R$ {despesas_transcorrer:,.2f}")
        
        with col_res3:
            st.metric("üìÖ Receitas a Transcorrer", f"R$ {receitas_transcorrer:,.2f}")
    else:
        st.info("üì≠ Nenhum lan√ßamento registrado ainda.")

# TAB 2: Visualizar Lan√ßamentos
with tab2:
    st.subheader("Hist√≥rico de Lan√ßamentos")
    
    df_lancamentos = carregar_lancamentos()
    
    # Corrigir tipos de dados no DataFrame
    for col in df_lancamentos.select_dtypes(include=['object']).columns:
        df_lancamentos[col] = df_lancamentos[col].astype(str)
    
    if not df_lancamentos.empty:
        # Filtros
        col_filt1, col_filt2, col_filt3 = st.columns(3)
        
        with col_filt1:
            tipo_filtro = st.multiselect(
                "Filtrar por Tipo",
                options=df_lancamentos['Tipo'].unique(),
                default=df_lancamentos['Tipo'].unique()
            )
        
        with col_filt2:
            categoria_filtro = st.multiselect(
                "Filtrar por Categoria",
                options=df_lancamentos['Categoria'].unique(),
                default=df_lancamentos['Categoria'].unique()
            )
        
        with col_filt3:
            data_inicio = st.date_input("Data Inicial", value=df_lancamentos['Data'].min())
            data_fim = st.date_input("Data Final", value=df_lancamentos['Data'].max())
        
        # Aplicar filtros
        df_filtrado = df_lancamentos[
            (df_lancamentos['Tipo'].isin(tipo_filtro)) &
            (df_lancamentos['Categoria'].isin(categoria_filtro)) &
            (df_lancamentos['Data'].dt.date >= data_inicio) &
            (df_lancamentos['Data'].dt.date <= data_fim)
        ]
        
        # Exibir tabela
        st.dataframe(df_filtrado, use_container_width=True, hide_index=True)
        
        # Estat√≠sticas
        st.divider()
        st.subheader("üìä Estat√≠sticas do Per√≠odo")
        
        col_stat1, col_stat2, col_stat3 = st.columns(3)
        
        receitas = df_filtrado[df_filtrado['Tipo'] == 'Receita']['Valor'].sum()
        despesas = df_filtrado[df_filtrado['Tipo'] == 'Despesa']['Valor'].sum()
        saldo = receitas - despesas
        
        with col_stat1:
            st.metric("üíö Total Receitas", f"R$ {receitas:,.2f}")
        
        with col_stat2:
            st.metric("‚ù§Ô∏è Total Despesas", f"R$ {despesas:,.2f}")
        
        with col_stat3:
            cor = "green" if saldo >= 0 else "red"
            st.metric("üíô Saldo", f"R$ {saldo:,.2f}", delta=None)
    else:
        st.info("üì≠ Nenhum lan√ßamento registrado ainda.")

# TAB 3: Resumo
with tab3:
    st.subheader("üìà Resumo Financeiro")
    
    df_lancamentos = carregar_lancamentos()
    
    if not df_lancamentos.empty:
        # Totais gerais
        col_total1, col_total2, col_total3 = st.columns(3)
        
        total_receitas = df_lancamentos[df_lancamentos['Tipo'] == 'Receita']['Valor'].sum()
        total_despesas = df_lancamentos[df_lancamentos['Tipo'] == 'Despesa']['Valor'].sum()
        saldo_geral = total_receitas - total_despesas
        
        with col_total1:
            st.metric("üíö Total Receitas", f"R$ {total_receitas:,.2f}")
        
        with col_total2:
            st.metric("‚ù§Ô∏è Total Despesas", f"R$ {total_despesas:,.2f}")
        
        with col_total3:
            st.metric("üíô Saldo Geral", f"R$ {saldo_geral:,.2f}")
        
        st.divider()
        
        # Gr√°ficos
        col_chart1, col_chart2 = st.columns(2)
        
        with col_chart1:
            st.subheader("Despesas por Categoria")
            despesas_cat = df_lancamentos[df_lancamentos['Tipo'] == 'Despesa'].groupby('Categoria')['Valor'].sum().sort_values(ascending=False)
            if not despesas_cat.empty:
                st.bar_chart(despesas_cat)
            else:
                st.info("Nenhuma despesa registrada.")
        
        with col_chart2:
            st.subheader("Receitas por Categoria")
            receitas_cat = df_lancamentos[df_lancamentos['Tipo'] == 'Receita'].groupby('Categoria')['Valor'].sum().sort_values(ascending=False)
            if not receitas_cat.empty:
                st.bar_chart(receitas_cat)
            else:
                st.info("Nenhuma receita registrada.")
        
        st.divider()
        
        # Resumo por m√©todo
        st.subheader("M√©todos de Pagamento Utilizados")
        metodos_uso = df_lancamentos.groupby('M√©todo')['Valor'].sum().sort_values(ascending=False)
        if not metodos_uso.empty:
            st.bar_chart(metodos_uso)
        else:
            st.info("Nenhum m√©todo registrado.")
    else:
        st.info("üì≠ Nenhum lan√ßamento registrado ainda. Comece adicionando um novo lan√ßamento!")

# Rodap√©
st.divider()
st.markdown("""
    <div style='text-align: center; color: #999; font-size: 12px; margin-top: 20px;'>
    üí∞ Aplicativo de Registro de Lan√ßamentos Financeiros | Desenvolvido com Python e Streamlit
    </div>
""", unsafe_allow_html=True)
